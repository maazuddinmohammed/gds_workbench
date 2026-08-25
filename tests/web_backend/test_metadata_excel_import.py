from collections.abc import Callable
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from gds_workbench_api.features.metadata.workbook import (
    MANIFEST_SHEET_NAME,
    MetadataWorkbookParseError,
    MetadataWorkbookSheet,
    build_metadata_workbook,
    parse_metadata_workbook,
)


def _workbook() -> bytes:
    return build_metadata_workbook(
        tenant_id=7,
        sheets=(
            MetadataWorkbookSheet(
                code="copy_group",
                name="Copy Groups",
                columns=(
                    "tenant_code",
                    "system_code",
                    "copy_group_name",
                    "copy_group_description",
                    "is_member_group_required",
                    "is_active",
                ),
                canonical_key=("tenant_code", "system_code", "copy_group_name"),
                row_schema={
                    "properties": {},
                },
                rows=(
                    {
                        "tenant_code": "NWA",
                        "system_code": "CRM",
                        "copy_group_name": "CRM daily",
                        "copy_group_description": "Daily customer load",
                        "is_member_group_required": False,
                        "is_active": True,
                    },
                ),
            ),
        ),
    )


def _canonical_workbook() -> bytes:
    from gds_etl_workbench.tools.snapshots.metadata.contracts import DATASETS_BY_NAME

    definition = DATASETS_BY_NAME["copy_group"]
    return build_metadata_workbook(
        tenant_id=7,
        sheets=(
            MetadataWorkbookSheet(
                code="copy_group",
                name=definition.label,
                columns=tuple(definition.row_model.model_fields),
                canonical_key=definition.canonical_key,
                row_schema=definition.row_model.model_json_schema(),
                rows=(
                    {
                        "tenant_code": "NWA",
                        "system_code": "CRM",
                        "copy_group_name": "CRM daily",
                        "copy_group_description": "Daily customer load",
                        "is_member_group_required": False,
                        "is_active": True,
                    },
                ),
            ),
        ),
    )


def _save_modified(content: bytes, change: Callable[[Workbook], object]) -> bytes:
    workbook = load_workbook(BytesIO(content), read_only=False, data_only=False)
    try:
        change(workbook)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()
    finally:
        workbook.close()


def test_canonical_metadata_workbook_round_trips_to_validated_pending_records() -> None:
    parsed = parse_metadata_workbook(_canonical_workbook(), tenant_id=7)

    assert len(parsed) == 1
    assert parsed[0].code == "copy_group"
    assert parsed[0].rows == (
        {
            "tenant_code": "NWA",
            "system_code": "CRM",
            "copy_group_name": "CRM daily",
            "copy_group_description": "Daily customer load",
            "is_member_group_required": False,
            "is_active": True,
        },
    )


def test_import_rejects_a_workbook_for_another_tenant() -> None:
    with pytest.raises(MetadataWorkbookParseError, match="manifest row is invalid"):
        parse_metadata_workbook(_canonical_workbook(), tenant_id=8)


def test_import_rejects_formula_cells_before_parsing_values() -> None:
    content = _save_modified(
        _canonical_workbook(),
        lambda workbook: setattr(workbook["Copy Groups"]["D2"], "value", "=NOW()"),
    )

    with pytest.raises(MetadataWorkbookParseError, match="unsafe content"):
        parse_metadata_workbook(content, tenant_id=7)


def test_import_rejects_extra_sheets_not_bound_by_the_manifest() -> None:
    content = _save_modified(
        _canonical_workbook(),
        lambda workbook: workbook.create_sheet("Notes"),
    )

    with pytest.raises(MetadataWorkbookParseError, match="unexpected sheet"):
        parse_metadata_workbook(content, tenant_id=7)


def test_import_rejects_a_changed_normalized_header() -> None:
    content = _save_modified(
        _canonical_workbook(),
        lambda workbook: setattr(workbook["Copy Groups"]["A1"], "value", "Tenant"),
    )

    with pytest.raises(MetadataWorkbookParseError, match="sheet header is invalid"):
        parse_metadata_workbook(content, tenant_id=7)


def test_import_rejects_macro_payloads_even_when_the_workbook_xml_is_valid() -> None:
    source = BytesIO(_canonical_workbook())
    output = BytesIO()
    with (
        ZipFile(source, mode="r") as current,
        ZipFile(output, mode="w", compression=ZIP_DEFLATED) as modified,
    ):
        for member in current.infolist():
            modified.writestr(member, current.read(member))
        modified.writestr("xl/vbaProject.bin", b"not-a-real-macro")

    with pytest.raises(MetadataWorkbookParseError, match="unsafe content"):
        parse_metadata_workbook(output.getvalue(), tenant_id=7)


def test_import_requires_the_hidden_manifest() -> None:
    content = _save_modified(
        _canonical_workbook(),
        lambda workbook: setattr(workbook[MANIFEST_SHEET_NAME], "sheet_state", "visible"),
    )

    with pytest.raises(MetadataWorkbookParseError, match="manifest visibility is invalid"):
        parse_metadata_workbook(content, tenant_id=7)


def test_manifest_schema_must_be_the_canonical_dataset_schema() -> None:
    with pytest.raises(MetadataWorkbookParseError, match="manifest contract is invalid"):
        parse_metadata_workbook(_workbook(), tenant_id=7)
