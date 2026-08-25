from __future__ import annotations

import json
from collections.abc import AsyncGenerator, Mapping, Sequence
from contextlib import asynccontextmanager
from datetime import date, datetime, timedelta, timezone
from hashlib import sha256
from io import BytesIO
from pathlib import Path
from typing import Any, LiteralString, Protocol, cast
from uuid import UUID, uuid4
from zipfile import ZipFile

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from gds_etl_workbench.adapters.auth.identity import AuthMode, IdentityProvider
from gds_etl_workbench.application.authorization import AuthorizationService
from gds_etl_workbench.domain.errors import WorkbenchError
from gds_etl_workbench.infrastructure.postgres import ReadIsolation, ReadTransaction
from gds_etl_workbench.tools.snapshots.metadata.contracts import DATASETS_BY_NAME
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from psycopg import Connection

from gds_workbench_api.database import WebPostgresDatabase
from gds_workbench_api.errors import workbench_error_response
from gds_workbench_api.features.metadata import (
    MAX_METADATA_EXPORT_ROWS_PER_SHEET,
    OPERATIONAL_DATASETS,
    DatabaseMetadataService,
    MetadataFilter,
    ObjectCatalogDetail,
    ObjectCatalogFilters,
    ObjectCatalogSummary,
    OperationalDataset,
    PostgresMetadataRepository,
    create_metadata_router,
)


class DisposablePostgres(Protocol):
    def connect_owner(self) -> Connection[dict[str, Any]]: ...

    def web_runtime_dsn(self) -> str: ...


DEMO_METADATA_SEED = (
    Path(__file__).parents[2] / "database" / "seed" / "01_metadata_snapshot_demo.sql"
)


class ExportTransaction:
    def __init__(
        self,
        *,
        expected_tenant_id: int = 7,
        authorized: bool = True,
    ) -> None:
        self.expected_tenant_id = expected_tenant_id
        self.authorized = authorized
        self.authorization_calls = 0
        self.queries: list[tuple[LiteralString, tuple[Any, ...]]] = []

    async def fetch_one(
        self,
        query: LiteralString,
        parameters: tuple[Any, ...] = (),
    ) -> dict[str, Any] | None:
        assert "WITH actor AS" in query
        assert "authorize_tenant_operation" not in query
        assert parameters[-1] == self.expected_tenant_id
        self.authorization_calls += 1
        return {
            "principal_id": 41,
            "principal_display_name": "Maaz",
            "is_super_admin": False,
            "effective_role": "viewer",
            "authorized": self.authorized,
            "denial_code": None if self.authorized else "authorization_denied",
            "lock_owner_display_name": None,
            "lock_expires_time": None,
        }

    async def fetch_all(
        self,
        query: LiteralString,
        parameters: tuple[Any, ...] = (),
    ) -> list[dict[str, Any]]:
        self.queries.append((query, parameters))
        return []


class ExportDatabase:
    def __init__(self, transaction: ExportTransaction | None = None) -> None:
        self.transaction = transaction or ExportTransaction()
        self.isolations: list[ReadIsolation] = []

    @asynccontextmanager
    async def read_transaction(
        self,
        *,
        isolation: ReadIsolation = ReadIsolation.READ_COMMITTED,
    ) -> AsyncGenerator[ExportTransaction]:
        self.isolations.append(isolation)
        yield self.transaction


class ExportRepository:
    def __init__(
        self,
        rows: Mapping[OperationalDataset, Sequence[Mapping[str, object]]] | None = None,
    ) -> None:
        self.export_calls: list[tuple[int, OperationalDataset, int]] = []
        self.rows = (
            {
                "copy_group": (
                    {
                        "tenant_code": "NWA",
                        "system_code": "CRM",
                        "copy_group_name": "CRM daily",
                        "copy_group_description": None,
                        "is_member_group_required": False,
                        "is_active": True,
                    },
                ),
                "process_group": (
                    {
                        "tenant_code": "NWA",
                        "system_code": "CRM",
                        "zone_code": "bronze",
                        "process_group_name": "CRM bronze",
                        "process_group_description": None,
                        "copy_group_name": "CRM daily",
                        "is_active": True,
                    },
                ),
            }
            if rows is None
            else rows
        )

    async def list_export_rows(
        self,
        transaction: ReadTransaction,
        *,
        tenant_id: int,
        dataset: OperationalDataset,
        limit: int,
    ) -> Sequence[Mapping[str, object]]:
        assert transaction is not None
        self.export_calls.append((tenant_id, dataset, limit))
        return self.rows.get(dataset, ())

    async def list_rows(
        self,
        transaction: ReadTransaction,
        *,
        tenant_id: int,
        dataset: OperationalDataset,
        filters: tuple[MetadataFilter, ...],
        limit: int,
        offset: int,
    ) -> Sequence[Mapping[str, object]]:
        raise AssertionError("not used")

    async def list_objects(
        self,
        transaction: ReadTransaction,
        *,
        tenant_id: int,
        filters: ObjectCatalogFilters,
        limit: int,
        offset: int,
    ) -> Sequence[ObjectCatalogSummary]:
        raise AssertionError("not used")

    async def get_object(
        self,
        transaction: ReadTransaction,
        *,
        tenant_id: int,
        object_id: int,
    ) -> ObjectCatalogDetail | None:
        raise AssertionError("not used")


def _identity_provider() -> IdentityProvider:
    return IdentityProvider(
        AuthMode.DEV,
        local_tenant_id=UUID("11111111-1111-1111-1111-111111111111"),
        local_principal_object_id=UUID("22222222-2222-2222-2222-222222222222"),
    )


def test_selected_operational_sheets_download_as_one_canonical_xlsx() -> None:
    database = ExportDatabase()
    repository = ExportRepository()
    service = DatabaseMetadataService(
        database=database,
        repository=cast(Any, repository),
        authorizer=AuthorizationService(),
        cursor_signing_key=b"development-only-key-32-bytes-long",
    )
    app = FastAPI()
    app.add_exception_handler(WorkbenchError, workbench_error_response)
    app.include_router(
        create_metadata_router(
            identity_provider=_identity_provider(),
            service=service,
        )
    )

    with TestClient(app) as client:
        response = client.post(
            "/api/v1/tenants/7/metadata/exports/xlsx",
            json={
                "schema_version": "1.0",
                "sheet_codes": ["process_group", "copy_group"],
            },
        )

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["content-disposition"] == (
        'attachment; filename="gds_operational_metadata__tenant_7__2_sheets.xlsx"'
    )
    assert response.headers["cache-control"] == "no-store"
    assert response.headers["x-content-type-options"] == "nosniff"
    workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=False)
    try:
        assert workbook.sheetnames == [
            "Copy Groups",
            "Process Groups",
            "__gds_manifest",
        ]
        assert tuple(
            cell.value for cell in next(workbook["Copy Groups"].iter_rows(max_row=1))
        ) == (
            "tenant_code",
            "system_code",
            "copy_group_name",
            "copy_group_description",
            "is_member_group_required",
            "is_active",
        )
        assert tuple(
            cell.value for cell in next(workbook["Copy Groups"].iter_rows(min_row=2))
        ) == (
            "NWA",
            "CRM",
            "CRM daily",
            None,
            False,
            True,
        )
    finally:
        workbook.close()
    assert database.isolations == [ReadIsolation.REPEATABLE_READ]
    assert [call[1] for call in repository.export_calls] == [
        "copy_group",
        "process_group",
    ]


def test_hidden_manifest_binds_each_selected_sheet_to_its_versioned_row_schema() -> (
    None
):
    service = DatabaseMetadataService(
        database=ExportDatabase(),
        repository=cast(Any, ExportRepository()),
        authorizer=AuthorizationService(),
        cursor_signing_key=b"development-only-key-32-bytes-long",
    )
    app = FastAPI()
    app.include_router(
        create_metadata_router(
            identity_provider=_identity_provider(),
            service=service,
        )
    )

    with TestClient(app) as client:
        response = client.post(
            "/api/v1/tenants/7/metadata/exports/xlsx",
            json={"schema_version": "1.0", "sheet_codes": ["copy_group"]},
        )

    workbook = load_workbook(
        BytesIO(response.content), read_only=False, data_only=False
    )
    try:
        manifest = workbook["__gds_manifest"]
        assert manifest.sheet_state == "veryHidden"
        manifest_rows = list(manifest.iter_rows(values_only=True))
        assert manifest_rows[0] == (
            "manifest_version",
            "workbook_kind",
            "tenant_id",
            "sheet_code",
            "sheet_name",
            "sheet_order",
            "row_count",
            "field_order_json",
            "canonical_key_json",
            "row_schema_json",
            "row_schema_sha256",
            "natural_key_normalization_json",
        )
        row = manifest_rows[1]
        assert row[:7] == (
            "1.0",
            "gds.operational_metadata",
            7,
            "copy_group",
            "Copy Groups",
            1,
            1,
        )
        fields = json.loads(cast(str, row[7]))
        canonical_key = json.loads(cast(str, row[8]))
        row_schema_json = cast(str, row[9])
        row_schema = json.loads(row_schema_json)
        assert fields == [
            "tenant_code",
            "system_code",
            "copy_group_name",
            "copy_group_description",
            "is_member_group_required",
            "is_active",
        ]
        assert canonical_key == ["tenant_code", "system_code", "copy_group_name"]
        assert list(row_schema["properties"]) == fields
        assert row[10] == sha256(row_schema_json.encode()).hexdigest()
        assert json.loads(cast(str, row[11])) == {
            "case": "unicode-lowercase",
            "other_values": "identity",
            "string_field_suffixes": ["_code", "_name", "_schema"],
            "trim_code_points": ["U+0020"],
            "unicode_normalization": "none",
            "version": "1.0",
        }
    finally:
        workbook.close()


def test_export_preserves_canonical_cell_types_and_writes_sql_as_literal_text() -> None:
    repository = ExportRepository(
        {
            "member_group": (
                {
                    "tenant_code": "NWA",
                    "system_code": "CRM",
                    "member_group_name": "Daily",
                    "member_group_description": None,
                    "member_group_initial_load_date": date(2026, 8, 24),
                    "is_active": True,
                },
            ),
            "copy_group_control": (
                {
                    "tenant_code": "NWA",
                    "system_code": "CRM",
                    "copy_group_name": "CRM daily",
                    "member_group_name": None,
                    "copy_group_control_initial_load_date": date(2026, 8, 23),
                    "copy_group_control_last_run_time": datetime(
                        2026,
                        8,
                        24,
                        14,
                        30,
                        tzinfo=timezone(timedelta(hours=2)),
                    ),
                    "copy_group_control_last_run_value": "@last-value",
                },
            ),
            "copy": (
                {
                    "tenant_code": "NWA",
                    "system_code": "CRM",
                    "copy_group_name": "CRM daily",
                    "source_tenant_code": "NWA",
                    "source_system_code": "CRM",
                    "source_connection_code": "MAIN",
                    "source_object_schema": "sales",
                    "source_object_name": "Customer",
                    "target_tenant_code": "NWA",
                    "target_system_code": "LAKE",
                    "target_connection_code": "MAIN",
                    "target_object_schema": "bronze",
                    "target_object_name": "Customer",
                    "copy_source_record_limit": "9223372036854775807",
                    "copy_source_record_limit_attribute": None,
                    "chunk_type_name": None,
                    "copy_source_initial_sql_script": "=SELECT * FROM sales.Customer",
                    "copy_source_incremental_sql_script": "@SELECT changed rows",
                    "copy_source_file_name": "+customer.csv",
                    "copy_source_file_pattern": "-customer-*.csv",
                    "copy_source_file_delimiter": ",",
                    "source_file_type_name": None,
                    "copy_source_order": 1,
                    "source_data_operation_name": "read",
                    "target_data_operation_name": "append",
                    "is_active": False,
                },
            ),
        }
    )
    service = DatabaseMetadataService(
        database=ExportDatabase(),
        repository=cast(Any, repository),
        authorizer=AuthorizationService(),
        cursor_signing_key=b"development-only-key-32-bytes-long",
    )
    app = FastAPI()
    app.include_router(
        create_metadata_router(identity_provider=_identity_provider(), service=service)
    )

    with TestClient(app) as client:
        response = client.post(
            "/api/v1/tenants/7/metadata/exports/xlsx",
            json={
                "schema_version": "1.0",
                "sheet_codes": ["copy", "copy_group_control", "member_group"],
            },
        )

    workbook = load_workbook(
        BytesIO(response.content), read_only=False, data_only=False
    )
    try:
        assert workbook.sheetnames == [
            "Member Groups",
            "Copy Group Controls",
            "Copies",
            "__gds_manifest",
        ]
        copy = workbook["Copies"]
        by_header: dict[str, Cell] = {}
        for header_cell in copy[1]:
            assert isinstance(header_cell.value, str)
            assert isinstance(header_cell.column, int)
            by_header[header_cell.value] = cast(
                Cell,
                copy.cell(row=2, column=header_cell.column),
            )
        assert by_header["copy_source_initial_sql_script"].value == (
            "=SELECT * FROM sales.Customer"
        )
        assert by_header["copy_source_initial_sql_script"].data_type == "s"
        assert by_header["copy_source_incremental_sql_script"].data_type == "s"
        assert by_header["copy_source_file_name"].data_type == "s"
        assert by_header["copy_source_file_pattern"].data_type == "s"
        assert by_header["copy_source_record_limit"].data_type == "s"
        assert by_header["copy_source_order"].data_type == "n"
        assert by_header["is_active"].data_type == "b"
        member_date = workbook["Member Groups"]["E2"]
        assert member_date.value == datetime(2026, 8, 24)
        assert member_date.data_type == "d"
        control_time = workbook["Copy Group Controls"]["F2"]
        assert control_time.value == datetime(2026, 8, 24, 12, 30)
        assert control_time.data_type == "d"
        copy_manifest_row = next(
            row
            for row in workbook["__gds_manifest"].iter_rows(values_only=True)
            if row[3] == "copy"
        )
        copy_manifest_fields = json.loads(cast(str, copy_manifest_row[7]))
        assert "copy_source_initial_sql_script" in copy_manifest_fields
        assert "copy_source_incremental_sql_script" in copy_manifest_fields
    finally:
        workbook.close()

    with ZipFile(BytesIO(response.content)) as archive:
        names = archive.namelist()
        assert all("vbaProject" not in name for name in names)
        assert all(not name.startswith("xl/externalLinks/") for name in names)
        worksheet_xml = b"".join(
            archive.read(name)
            for name in names
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        )
        relationship_xml = b"".join(
            archive.read(name) for name in names if name.endswith(".rels")
        )
        assert b"<f" not in worksheet_xml
        assert b'TargetMode="External"' not in relationship_xml


def test_identical_export_inputs_produce_byte_identical_canonical_xlsx_packages() -> (
    None
):
    service = DatabaseMetadataService(
        database=ExportDatabase(),
        repository=cast(Any, ExportRepository()),
        authorizer=AuthorizationService(),
        cursor_signing_key=b"development-only-key-32-bytes-long",
    )
    app = FastAPI()
    app.include_router(
        create_metadata_router(identity_provider=_identity_provider(), service=service)
    )
    request = {"schema_version": "1.0", "sheet_codes": ["copy_group"]}

    with TestClient(app) as client:
        first = client.post("/api/v1/tenants/7/metadata/exports/xlsx", json=request)
        second = client.post("/api/v1/tenants/7/metadata/exports/xlsx", json=request)

    assert first.content == second.content
    with ZipFile(BytesIO(first.content)) as archive:
        assert archive.namelist() == sorted(archive.namelist())
        assert {member.date_time for member in archive.infolist()} == {
            (1980, 1, 1, 0, 0, 0)
        }


def test_sheet_rows_are_sorted_by_the_normalized_canonical_key() -> None:
    repository = ExportRepository(
        {
            "copy_group": (
                {
                    "tenant_code": "NWA",
                    "system_code": "CRM",
                    "copy_group_name": "Zulu",
                    "copy_group_description": None,
                    "is_member_group_required": False,
                    "is_active": True,
                },
                {
                    "tenant_code": "NWA",
                    "system_code": "CRM",
                    "copy_group_name": " alpha ",
                    "copy_group_description": None,
                    "is_member_group_required": False,
                    "is_active": True,
                },
            )
        }
    )
    service = DatabaseMetadataService(
        database=ExportDatabase(),
        repository=cast(Any, repository),
        authorizer=AuthorizationService(),
        cursor_signing_key=b"development-only-key-32-bytes-long",
    )
    app = FastAPI()
    app.include_router(
        create_metadata_router(identity_provider=_identity_provider(), service=service)
    )

    with TestClient(app) as client:
        response = client.post(
            "/api/v1/tenants/7/metadata/exports/xlsx",
            json={"schema_version": "1.0", "sheet_codes": ["copy_group"]},
        )

    workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=False)
    try:
        names = [
            row[2].value
            for row in workbook["Copy Groups"].iter_rows(min_row=2, max_col=3)
        ]
        assert names == [" alpha ", "Zulu"]
    finally:
        workbook.close()


def test_all_selection_exports_every_operational_sheet_even_when_empty() -> None:
    repository = ExportRepository({})
    service = DatabaseMetadataService(
        database=ExportDatabase(),
        repository=cast(Any, repository),
        authorizer=AuthorizationService(),
        cursor_signing_key=b"development-only-key-32-bytes-long",
    )
    app = FastAPI()
    app.include_router(
        create_metadata_router(identity_provider=_identity_provider(), service=service)
    )

    with TestClient(app) as client:
        response = client.post(
            "/api/v1/tenants/7/metadata/exports/xlsx",
            json={"schema_version": "1.0", "sheet_codes": "all"},
        )

    workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=False)
    try:
        assert workbook.sheetnames == [
            "Source Objects",
            "Source Attributes",
            "Bronze Objects",
            "Bronze Attributes",
            "Silver Objects",
            "Silver Attributes",
            "Gold Objects",
            "Gold Attributes",
            "Ingestion Object Mappings",
            "Ingestion Attribute Mappings",
            "Copy Groups",
            "Member Groups",
            "Copy Group Controls",
            "Copies",
            "Process Groups",
            "Processes",
            "__gds_manifest",
        ]
        assert all(workbook[sheet].max_row == 1 for sheet in workbook.sheetnames[:-1])
        for code, sheet_name in zip(
            OPERATIONAL_DATASETS,
            workbook.sheetnames[:-1],
            strict=True,
        ):
            assert tuple(cell.value for cell in workbook[sheet_name][1]) == tuple(
                DATASETS_BY_NAME[code].row_model.model_fields
            )
        assert workbook["__gds_manifest"].max_row == 17
    finally:
        workbook.close()
    assert len(repository.export_calls) == 16


def test_export_selection_rejects_empty_duplicate_and_nonoperational_sheet_codes() -> (
    None
):
    database = ExportDatabase()
    repository = ExportRepository({})
    service = DatabaseMetadataService(
        database=database,
        repository=cast(Any, repository),
        authorizer=AuthorizationService(),
        cursor_signing_key=b"development-only-key-32-bytes-long",
    )
    app = FastAPI()
    app.add_exception_handler(WorkbenchError, workbench_error_response)
    app.include_router(
        create_metadata_router(identity_provider=_identity_provider(), service=service)
    )

    with TestClient(app) as client:
        responses = [
            client.post(
                "/api/v1/tenants/7/metadata/exports/xlsx",
                json={"schema_version": "1.0", "sheet_codes": sheet_codes},
            )
            for sheet_codes in (
                [],
                ["copy", "copy"],
                ["project"],
                ["connection_value"],
                ["x" * 101],
            )
        ]

    assert [response.status_code for response in responses] == [422, 422, 422, 422, 422]
    assert all(
        response.json()["error"]["code"] == "invalid_request"
        and response.json()["error"]["message"]
        == "Metadata workbook sheet selection is invalid."
        for response in responses
    )
    assert all(
        forbidden not in response.text
        for response in responses
        for forbidden in ("project", "connection_value")
    )
    assert database.isolations == []
    assert repository.export_calls == []


def test_unauthorized_tenant_export_is_denied_before_any_metadata_read() -> None:
    transaction = ExportTransaction(expected_tenant_id=8, authorized=False)
    database = ExportDatabase(transaction)
    repository = ExportRepository({})
    service = DatabaseMetadataService(
        database=database,
        repository=cast(Any, repository),
        authorizer=AuthorizationService(),
        cursor_signing_key=b"development-only-key-32-bytes-long",
    )
    app = FastAPI()
    app.add_exception_handler(WorkbenchError, workbench_error_response)
    app.include_router(
        create_metadata_router(identity_provider=_identity_provider(), service=service)
    )

    with TestClient(app) as client:
        response = client.post(
            "/api/v1/tenants/8/metadata/exports/xlsx",
            json={"schema_version": "1.0", "sheet_codes": ["copy_group"]},
        )

    assert response.status_code == 403
    assert response.json()["error"]["code"] == "authorization_denied"
    assert repository.export_calls == []
    assert transaction.authorization_calls == 1
    assert database.isolations == [ReadIsolation.REPEATABLE_READ]


def test_export_rejects_an_oversized_excel_cell_without_echoing_its_value() -> None:
    oversized_value = "SENSITIVE_VALUE_" + ("x" * 32_768)
    repository = ExportRepository(
        {
            "copy_group": (
                {
                    "tenant_code": "NWA",
                    "system_code": "CRM",
                    "copy_group_name": "CRM daily",
                    "copy_group_description": oversized_value,
                    "is_member_group_required": False,
                    "is_active": True,
                },
            )
        }
    )
    service = DatabaseMetadataService(
        database=ExportDatabase(),
        repository=cast(Any, repository),
        authorizer=AuthorizationService(),
        cursor_signing_key=b"development-only-key-32-bytes-long",
    )
    app = FastAPI()
    app.add_exception_handler(WorkbenchError, workbench_error_response)
    app.include_router(
        create_metadata_router(identity_provider=_identity_provider(), service=service)
    )

    with TestClient(app) as client:
        response = client.post(
            "/api/v1/tenants/7/metadata/exports/xlsx",
            json={"schema_version": "1.0", "sheet_codes": ["copy_group"]},
        )

    assert response.status_code == 422
    assert response.json()["error"]["code"] == "invalid_request"
    assert response.json()["error"]["message"] == (
        "Metadata workbook contains a value that Excel cannot represent."
    )
    assert "SENSITIVE_VALUE" not in response.text


def test_export_redacts_an_illegal_xml_cell_failure() -> None:
    repository = ExportRepository(
        {
            "copy_group": (
                {
                    "tenant_code": "NWA",
                    "system_code": "CRM",
                    "copy_group_name": "CRM daily",
                    "copy_group_description": "SECRET\x00VALUE",
                    "is_member_group_required": False,
                    "is_active": True,
                },
            )
        }
    )
    service = DatabaseMetadataService(
        database=ExportDatabase(),
        repository=cast(Any, repository),
        authorizer=AuthorizationService(),
        cursor_signing_key=b"development-only-key-32-bytes-long",
    )
    app = FastAPI()
    app.add_exception_handler(WorkbenchError, workbench_error_response)
    app.include_router(
        create_metadata_router(identity_provider=_identity_provider(), service=service)
    )

    with TestClient(app) as client:
        response = client.post(
            "/api/v1/tenants/7/metadata/exports/xlsx",
            json={"schema_version": "1.0", "sheet_codes": ["copy_group"]},
        )

    assert response.status_code == 422
    assert response.json()["error"]["message"] == (
        "Metadata workbook contains a value that Excel cannot represent."
    )
    assert "SECRET" not in response.text


def test_export_redacts_a_canonical_row_validation_failure() -> None:
    repository = ExportRepository(
        {
            "copy_group": (
                {
                    "tenant_code": "NWA",
                    "system_code": "CRM",
                    "copy_group_name": "SENSITIVE_INVALID_" + ("x" * 201),
                    "copy_group_description": None,
                    "is_member_group_required": False,
                    "is_active": True,
                },
            )
        }
    )
    service = DatabaseMetadataService(
        database=ExportDatabase(),
        repository=cast(Any, repository),
        authorizer=AuthorizationService(),
        cursor_signing_key=b"development-only-key-32-bytes-long",
    )
    app = FastAPI()
    app.add_exception_handler(WorkbenchError, workbench_error_response)
    app.include_router(
        create_metadata_router(identity_provider=_identity_provider(), service=service)
    )

    with TestClient(app, raise_server_exceptions=False) as client:
        response = client.post(
            "/api/v1/tenants/7/metadata/exports/xlsx",
            json={"schema_version": "1.0", "sheet_codes": ["copy_group"]},
        )

    assert response.status_code == 422
    assert response.json()["error"]["message"] == (
        "Metadata workbook row does not match its canonical schema."
    )
    assert "SENSITIVE_INVALID" not in response.text


class OversizedExportRepository(ExportRepository):
    async def list_export_rows(
        self,
        transaction: ReadTransaction,
        *,
        tenant_id: int,
        dataset: OperationalDataset,
        limit: int,
    ) -> Sequence[Mapping[str, object]]:
        row = cast(Mapping[str, object], ExportRepository().rows["copy_group"][0])
        self.export_calls.append((tenant_id, dataset, limit))
        return [row] * (MAX_METADATA_EXPORT_ROWS_PER_SHEET + 1)


def test_export_rejects_a_sheet_above_the_fixed_row_limit() -> None:
    repository = OversizedExportRepository({})
    service = DatabaseMetadataService(
        database=ExportDatabase(),
        repository=cast(Any, repository),
        authorizer=AuthorizationService(),
        cursor_signing_key=b"development-only-key-32-bytes-long",
    )
    app = FastAPI()
    app.add_exception_handler(WorkbenchError, workbench_error_response)
    app.include_router(
        create_metadata_router(identity_provider=_identity_provider(), service=service)
    )

    with TestClient(app) as client:
        response = client.post(
            "/api/v1/tenants/7/metadata/exports/xlsx",
            json={"schema_version": "1.0", "sheet_codes": ["copy_group"]},
        )

    assert response.status_code == 422
    assert response.json()["error"]["message"] == (
        "Metadata workbook exceeds its row limit."
    )
    assert repository.export_calls == [
        (7, "copy_group", MAX_METADATA_EXPORT_ROWS_PER_SHEET + 1)
    ]


@pytest.mark.asyncio
async def test_export_repository_uses_one_fixed_tenant_scoped_bounded_query() -> None:
    transaction = ExportTransaction()
    repository = PostgresMetadataRepository()

    rows = await repository.list_export_rows(
        transaction,
        tenant_id=7,
        dataset="source_object",
        limit=MAX_METADATA_EXPORT_ROWS_PER_SHEET + 1,
    )

    assert rows == []
    assert len(transaction.queries) == 1
    query, parameters = transaction.queries[0]
    assert "WITH RECURSIVE requested_tenant AS" in query
    assert "connection_value" not in query
    assert "ORDER BY object.object_id" in query
    assert "LIMIT %s" in query
    assert "OFFSET" not in query
    assert parameters == (7, "source", MAX_METADATA_EXPORT_ROWS_PER_SHEET + 1)


def test_database_route_exports_authorized_tenant_rows_from_disposable_postgres(
    web_postgres_database: DisposablePostgres,
) -> None:
    suffix = uuid4().hex[:12]
    entra_tenant_id = uuid4()
    entra_object_id = uuid4()
    with web_postgres_database.connect_owner() as connection:
        existing = connection.execute(
            "SELECT tenant_id FROM core.tenant WHERE tenant_code = 'DEMO_TENANT'"
        ).fetchone()
        if existing is None:
            connection.execute(
                cast(LiteralString, DEMO_METADATA_SEED.read_text(encoding="utf-8"))
            )
        tenant = connection.execute(
            "SELECT tenant_id FROM core.tenant WHERE tenant_code = 'DEMO_TENANT'"
        ).fetchone()
        assert tenant is not None
        tenant_id = tenant["tenant_id"]
        assert isinstance(tenant_id, int) and not isinstance(tenant_id, bool)
        principal = connection.execute(
            """
            INSERT INTO security.principal (
                principal_type,
                principal_display_name,
                principal_email
            ) VALUES ('user', %s, %s)
            RETURNING principal_id
            """,
            (
                f"Metadata export viewer {suffix}",
                f"metadata_export_{suffix}@example.test",
            ),
        ).fetchone()
        assert principal is not None
        principal_id = principal["principal_id"]
        connection.execute(
            """
            INSERT INTO security.entra_principal_identity (
                principal_id,
                principal_type,
                entra_tenant_id,
                entra_object_id
            ) VALUES (%s, 'user', %s, %s)
            """,
            (principal_id, entra_tenant_id, entra_object_id),
        )
        connection.execute(
            """
            INSERT INTO security.tenant_principal_access (
                tenant_id,
                principal_id,
                tenant_role,
                granted_by_principal_id
            ) VALUES (%s, %s, 'viewer', %s)
            """,
            (tenant_id, principal_id, principal_id),
        )

    database = WebPostgresDatabase(
        dsn=web_postgres_database.web_runtime_dsn(),
        pool_min=1,
        pool_max=1,
        pool_timeout_seconds=5,
    )
    service = DatabaseMetadataService(
        database=database,
        repository=PostgresMetadataRepository(),
        authorizer=AuthorizationService(),
        cursor_signing_key=b"development-only-key-32-bytes-long",
    )

    @asynccontextmanager
    async def lifespan(_app: FastAPI) -> AsyncGenerator[None]:
        await database.open()
        try:
            yield
        finally:
            await database.close()

    app = FastAPI(lifespan=lifespan)
    app.add_exception_handler(WorkbenchError, workbench_error_response)
    app.include_router(
        create_metadata_router(
            identity_provider=IdentityProvider(
                AuthMode.DEV,
                local_tenant_id=entra_tenant_id,
                local_principal_object_id=entra_object_id,
            ),
            service=service,
        )
    )

    with TestClient(app) as client:
        response = client.post(
            f"/api/v1/tenants/{tenant_id}/metadata/exports/xlsx",
            json={
                "schema_version": "1.0",
                "sheet_codes": ["source_object", "copy"],
            },
        )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=False)
    try:
        assert workbook.sheetnames == ["Source Objects", "Copies", "__gds_manifest"]
        source_headers = tuple(cell.value for cell in workbook["Source Objects"][1])
        assert source_headers == tuple(
            DATASETS_BY_NAME["source_object"].row_model.model_fields
        )
        source_tenant_codes = {
            cast(str, row[0].value)
            for row in workbook["Source Objects"].iter_rows(min_row=2)
        }
        assert source_tenant_codes == {"DEMO_TENANT"}
        copy_headers = tuple(cell.value for cell in workbook["Copies"][1])
        assert "copy_source_initial_sql_script" in copy_headers
        assert "copy_source_incremental_sql_script" in copy_headers
        assert all("connection_value" not in str(cell) for cell in copy_headers)
    finally:
        workbook.close()
