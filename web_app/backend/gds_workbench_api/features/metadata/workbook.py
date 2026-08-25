"""In-memory XLSX rendering for governed operational metadata."""

import json
from collections.abc import Mapping
from dataclasses import dataclass
from datetime import UTC, date, datetime
from hashlib import sha256
from io import BytesIO
from zipfile import ZIP_DEFLATED, BadZipFile, ZipFile, ZipInfo

from gds_etl_workbench.tools.snapshots.metadata.contracts import (
    DATASETS_BY_NAME,
    natural_key_normalization_document,
)
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MANIFEST_SHEET_NAME = "__gds_manifest"
MAX_XLSX_CELL_CHARACTERS = 32_767
MAX_XLSX_TEXT_CHARACTERS = 16 * 1024 * 1024
MAX_XLSX_BYTES = 32 * 1024 * 1024
MAX_XLSX_PACKAGE_ENTRIES = 2_048
MAX_XLSX_UNCOMPRESSED_BYTES = 96 * 1024 * 1024
MAX_XLSX_ROWS_PER_SHEET = 10_000
MAX_XLSX_ROWS = 50_000

MANIFEST_COLUMNS = (
    "manifest_version",
    "workbook_kind",
    "tenant_id",
    "sheet_code",
    "sheet_name",
    "sheet_order",
    "row_count",
    "field_order_json",
    "canonical_key_json",
    "row_schema_json",
    "row_schema_sha256",
    "natural_key_normalization_json",
)


class MetadataWorkbookBuildError(ValueError):
    """One safe rendering failure with no cell content in its message."""


class MetadataWorkbookParseError(ValueError):
    """One safe import failure that never includes workbook cell content."""


@dataclass(frozen=True, slots=True)
class MetadataWorkbookSheet:
    code: str
    name: str
    columns: tuple[str, ...]
    canonical_key: tuple[str, ...]
    row_schema: Mapping[str, object]
    rows: tuple[Mapping[str, object], ...]


def build_metadata_workbook(
    *,
    tenant_id: int,
    sheets: tuple[MetadataWorkbookSheet, ...],
) -> bytes:
    """Render one bounded workbook without touching the filesystem."""
    text_characters = 0
    for definition in sheets:
        for row in definition.rows:
            for column in definition.columns:
                value = _validated_cell_value(row[column])
                if isinstance(value, str):
                    text_characters += len(value)
                    if text_characters > MAX_XLSX_TEXT_CHARACTERS:
                        raise MetadataWorkbookBuildError("XLSX text content is too large")

    workbook = Workbook()
    active_sheet = workbook.active
    if active_sheet is None:
        raise MetadataWorkbookBuildError("XLSX workbook initialization failed")
    workbook.remove(active_sheet)
    fixed_time = datetime(2000, 1, 1, tzinfo=UTC)
    workbook.properties.creator = "GDS Workbench"
    workbook.properties.created = fixed_time
    workbook.properties.modified = fixed_time

    for definition in sheets:
        worksheet = workbook.create_sheet(definition.name)
        worksheet.append(definition.columns)
        for row in definition.rows:
            worksheet.append([row[column] for column in definition.columns])
        for row in worksheet.iter_rows():
            for cell in row:
                _set_literal_cell_type(cell)

    manifest = workbook.create_sheet(MANIFEST_SHEET_NAME)
    manifest.append(MANIFEST_COLUMNS)
    normalization_json = json.dumps(
        natural_key_normalization_document(),
        ensure_ascii=False,
        separators=(",", ":"),
        sort_keys=True,
    )
    for sheet_order, definition in enumerate(sheets, start=1):
        row_schema_json = json.dumps(
            definition.row_schema,
            ensure_ascii=False,
            separators=(",", ":"),
        )
        manifest.append(
            (
                "1.0",
                "gds.operational_metadata",
                tenant_id,
                definition.code,
                definition.name,
                sheet_order,
                len(definition.rows),
                json.dumps(definition.columns, separators=(",", ":")),
                json.dumps(definition.canonical_key, separators=(",", ":")),
                row_schema_json,
                sha256(row_schema_json.encode()).hexdigest(),
                normalization_json,
            )
        )
    for row in manifest.iter_rows():
        for cell in row:
            _set_literal_cell_type(cell)
    manifest.sheet_state = "veryHidden"

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    content = _canonicalize_xlsx_package(buffer.getvalue())
    if len(content) > MAX_XLSX_BYTES:
        raise MetadataWorkbookBuildError("XLSX package is too large")
    return content


def parse_metadata_workbook(
    content: bytes,
    *,
    tenant_id: int,
) -> tuple[MetadataWorkbookSheet, ...]:
    """Validate and parse one canonical, in-memory operational workbook."""
    if not content or len(content) > MAX_XLSX_BYTES:
        raise MetadataWorkbookParseError("XLSX package size is invalid")
    _inspect_xlsx_package(content)

    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=False,
            data_only=False,
            keep_links=False,
        )
    except (BadZipFile, KeyError, OSError, ValueError) as error:
        raise MetadataWorkbookParseError("XLSX package could not be read") from error

    try:
        if MANIFEST_SHEET_NAME not in workbook.sheetnames:
            raise MetadataWorkbookParseError("XLSX manifest is missing")
        manifest = workbook[MANIFEST_SHEET_NAME]
        if manifest.sheet_state != "veryHidden":
            raise MetadataWorkbookParseError("XLSX manifest visibility is invalid")
        manifest_rows = list(manifest.iter_rows(values_only=True))
        if not manifest_rows or tuple(manifest_rows[0]) != MANIFEST_COLUMNS:
            raise MetadataWorkbookParseError("XLSX manifest header is invalid")
        if not 1 <= len(manifest_rows) - 1 <= len(DATASETS_BY_NAME):
            raise MetadataWorkbookParseError("XLSX manifest sheet count is invalid")

        normalization = natural_key_normalization_document()
        parsed: list[MetadataWorkbookSheet] = []
        expected_sheet_names: list[str] = []
        seen_codes: set[str] = set()
        total_rows = 0
        for expected_order, raw_manifest_row in enumerate(manifest_rows[1:], start=1):
            if len(raw_manifest_row) != len(MANIFEST_COLUMNS):
                raise MetadataWorkbookParseError("XLSX manifest row is invalid")
            (
                manifest_version,
                workbook_kind,
                manifest_tenant_id,
                sheet_code,
                sheet_name,
                sheet_order,
                row_count,
                field_order_json,
                canonical_key_json,
                row_schema_json,
                row_schema_sha256,
                normalization_json,
            ) = raw_manifest_row
            if (
                manifest_version != "1.0"
                or workbook_kind != "gds.operational_metadata"
                or manifest_tenant_id != tenant_id
                or type(sheet_order) is not int
                or sheet_order != expected_order
                or type(row_count) is not int
                or not 0 <= row_count <= MAX_XLSX_ROWS_PER_SHEET
                or not isinstance(sheet_code, str)
                or sheet_code not in DATASETS_BY_NAME
                or sheet_code in seen_codes
                or not isinstance(sheet_name, str)
                or not isinstance(field_order_json, str)
                or not isinstance(canonical_key_json, str)
                or not isinstance(row_schema_json, str)
                or not isinstance(row_schema_sha256, str)
                or not isinstance(normalization_json, str)
            ):
                raise MetadataWorkbookParseError("XLSX manifest row is invalid")

            definition = DATASETS_BY_NAME[sheet_code]
            columns = tuple(definition.row_model.model_fields)
            try:
                declared_columns = tuple(json.loads(field_order_json))
                declared_key = tuple(json.loads(canonical_key_json))
                declared_schema = json.loads(row_schema_json)
                declared_normalization = json.loads(normalization_json)
            except (json.JSONDecodeError, TypeError) as error:
                raise MetadataWorkbookParseError("XLSX manifest JSON is invalid") from error
            if (
                sheet_name != definition.label
                or declared_columns != columns
                or declared_key != definition.canonical_key
                or declared_schema != definition.row_model.model_json_schema()
                or row_schema_sha256 != sha256(row_schema_json.encode()).hexdigest()
                or declared_normalization != normalization
                or sheet_name not in workbook.sheetnames
            ):
                raise MetadataWorkbookParseError("XLSX manifest contract is invalid")

            worksheet = workbook[sheet_name]
            records = _parse_sheet_rows(
                worksheet,
                columns=columns,
                row_count=row_count,
                dataset=sheet_code,
            )
            total_rows += len(records)
            if total_rows > MAX_XLSX_ROWS:
                raise MetadataWorkbookParseError("XLSX row count is too large")
            parsed.append(
                MetadataWorkbookSheet(
                    code=sheet_code,
                    name=sheet_name,
                    columns=columns,
                    canonical_key=definition.canonical_key,
                    row_schema=definition.row_model.model_json_schema(),
                    rows=records,
                )
            )
            expected_sheet_names.append(sheet_name)
            seen_codes.add(sheet_code)

        if workbook.sheetnames != [*expected_sheet_names, MANIFEST_SHEET_NAME]:
            raise MetadataWorkbookParseError("XLSX contains an unexpected sheet")
        return tuple(parsed)
    finally:
        workbook.close()


def _inspect_xlsx_package(content: bytes) -> None:
    try:
        with ZipFile(BytesIO(content), mode="r") as package:
            members = package.infolist()
            if not members or len(members) > MAX_XLSX_PACKAGE_ENTRIES:
                raise MetadataWorkbookParseError("XLSX package structure is invalid")
            names = [member.filename for member in members]
            if len(names) != len(set(names)):
                raise MetadataWorkbookParseError("XLSX package contains duplicate entries")
            total_uncompressed = 0
            for member in members:
                normalized = member.filename.replace("\\", "/")
                if (
                    member.flag_bits & 0x1
                    or normalized.startswith("/")
                    or any(part in {"", ".", ".."} for part in normalized.split("/"))
                ):
                    raise MetadataWorkbookParseError("XLSX package structure is invalid")
                total_uncompressed += member.file_size
                if total_uncompressed > MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise MetadataWorkbookParseError("XLSX package is too large")
                lower_name = normalized.lower()
                payload = package.read(member)
                lower_payload = payload.lower()
                if (
                    "vbaproject" in lower_name
                    or lower_name.startswith("xl/externallinks/")
                    or lower_name == "xl/calcchain.xml"
                    or (
                        lower_name.startswith("xl/worksheets/")
                        and lower_name.endswith(".xml")
                        and b"<f" in lower_payload
                    )
                    or (lower_name.endswith(".rels") and b'targetmode="external"' in lower_payload)
                    or (lower_name == "[content_types].xml" and b"macroenabled" in lower_payload)
                ):
                    raise MetadataWorkbookParseError("XLSX package contains unsafe content")
    except BadZipFile as error:
        raise MetadataWorkbookParseError("XLSX package could not be read") from error


def _parse_sheet_rows(
    worksheet: Worksheet,
    *,
    columns: tuple[str, ...],
    row_count: int,
    dataset: str,
) -> tuple[Mapping[str, object], ...]:
    raw_rows = list(worksheet.iter_rows())
    if not raw_rows or tuple(cell.value for cell in raw_rows[0]) != columns:
        raise MetadataWorkbookParseError("XLSX sheet header is invalid")
    if len(raw_rows) - 1 != row_count:
        raise MetadataWorkbookParseError("XLSX sheet row count does not match its manifest")

    definition = DATASETS_BY_NAME[dataset]
    records: list[Mapping[str, object]] = []
    text_characters = 0
    for raw_row in raw_rows[1:]:
        if len(raw_row) != len(columns) or any(cell.data_type == "f" for cell in raw_row):
            raise MetadataWorkbookParseError("XLSX sheet row is invalid")
        values: dict[str, object] = {}
        for column, cell in zip(columns, raw_row, strict=True):
            value = cell.value
            if isinstance(value, str):
                text_characters += len(value)
                if len(value) > MAX_XLSX_CELL_CHARACTERS:
                    raise MetadataWorkbookParseError("XLSX cell value is too long")
            if value is not None and not isinstance(value, (str, bool, int, date, datetime)):
                raise MetadataWorkbookParseError("XLSX cell value type is unsupported")
            values[column] = value
        if text_characters > MAX_XLSX_TEXT_CHARACTERS:
            raise MetadataWorkbookParseError("XLSX text content is too large")
        try:
            record = definition.row_model.model_validate(values, strict=True)
        except ValidationError as error:
            raise MetadataWorkbookParseError(
                "XLSX row does not match its canonical schema"
            ) from error
        records.append(record.model_dump(mode="json"))

    for field_name, expected_value in definition.fixed_values:
        if any(record[field_name] != expected_value for record in records):
            raise MetadataWorkbookParseError("XLSX fixed metadata value is invalid")
    return tuple(records)


def _set_literal_cell_type(cell: Cell) -> None:
    value = cell.value
    if isinstance(value, str):
        cell.data_type = "s"
    elif isinstance(value, datetime):
        if value.tzinfo is not None:
            cell.value = value.astimezone(UTC).replace(tzinfo=None)
        cell.number_format = 'yyyy-mm-dd hh:mm:ss.000000"Z"'
    elif isinstance(value, date):
        cell.number_format = "yyyy-mm-dd"


def _validated_cell_value(value: object) -> object:
    if isinstance(value, str) and len(value) > MAX_XLSX_CELL_CHARACTERS:
        raise MetadataWorkbookBuildError("XLSX cell value is too long")
    if value is not None and not isinstance(value, (str, bool, int, date, datetime)):
        raise MetadataWorkbookBuildError("XLSX cell value type is unsupported")
    return value


def _canonicalize_xlsx_package(content: bytes) -> bytes:
    source = BytesIO(content)
    output = BytesIO()
    with (
        ZipFile(source, mode="r") as package,
        ZipFile(
            output,
            mode="w",
            compression=ZIP_DEFLATED,
            compresslevel=9,
            allowZip64=False,
        ) as canonical,
    ):
        for name in sorted(package.namelist()):
            payload = package.read(name)
            lower_name = name.lower()
            if (
                "vbaproject" in lower_name
                or lower_name.startswith("xl/externallinks/")
                or (
                    lower_name.startswith("xl/worksheets/")
                    and lower_name.endswith(".xml")
                    and b"<f" in payload
                )
                or (lower_name.endswith(".rels") and b'TargetMode="External"' in payload)
            ):
                raise MetadataWorkbookBuildError("XLSX package is unsafe")
            member = ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            member.compress_type = ZIP_DEFLATED
            member.create_system = 3
            member.external_attr = 0o100644 << 16
            canonical.writestr(member, payload)
    return output.getvalue()
