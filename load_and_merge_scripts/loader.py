"""Manual, fail-fast Excel loader for current GDS PostgreSQL tables."""

from __future__ import annotations

import argparse
import importlib
import os
import re
import sys
from collections.abc import Callable, Iterable, Mapping, Sequence
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Protocol, cast
from uuid import UUID

import psycopg
from psycopg.conninfo import conninfo_to_dict


class _YamlModule(Protocol):
    YAMLError: type[Exception]

    def safe_load(self, stream: str) -> object: ...


class _DotenvModule(Protocol):
    load_dotenv: Callable[..., bool]


SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR = SCRIPT_DIR / "data"
CONFIG_PATH = SCRIPT_DIR / "load_config.yaml"
ENV_PATH = SCRIPT_DIR / ".env"

WRITE_ACKNOWLEDGEMENT = "I_UNDERSTAND_THIS_WRITES_DATA"
LOCK_WORKBOOK = "locks.xlsx"
LOCK_SHEET = "LockControl"
LOCK_COLUMNS = (
    "schema_name",
    "table_name",
    "id_column_name",
    "id_value",
    "is_locked",
    "expected_model_revision",
)
LOCK_TARGETS = {
    ("core", "object", "object_id"): "is_locked",
    ("model", "model_scope", "model_scope_id"): "model_scope_is_locked",
    (
        "model",
        "modeling_assertion_record",
        "modeling_assertion_record_id",
    ): "modeling_assertion_record_is_locked",
}
BOOTSTRAP_ACTIVITY_TABLES = (
    ("security", "tenant_lock"),
    ("security", "tenant_lock_event"),
    ("model", "model_event_log"),
    ("model", "model_revision_transaction"),
    ("workflow", "attribute_profile"),
    ("workflow", "analysis_result"),
    ("workflow", "conceptual_object"),
    ("workflow", "conceptual_relationship"),
    ("workflow", "conceptual_support"),
    ("workflow", "logical_submodel"),
    ("workflow", "logical_entity"),
    ("workflow", "logical_entity_submodel"),
    ("workflow", "logical_attribute"),
    ("workflow", "logical_entity_source_mapping"),
    ("workflow", "logical_attribute_source_mapping"),
    ("workflow", "logical_relationship"),
    ("workflow", "dimensional_submodel"),
    ("workflow", "dimensional_entity"),
    ("workflow", "dimensional_entity_submodel"),
    ("workflow", "dimensional_attribute"),
    ("workflow", "dimensional_entity_source_mapping"),
    ("workflow", "dimensional_attribute_source_mapping"),
    ("workflow", "dimensional_relationship"),
    ("workflow", "mapping_source_system_dependency"),
    ("workflow", "mapping_object"),
    ("workflow", "mapping_attribute"),
    ("application", "workflow_run"),
    ("mcp", "model_change_set"),
    ("mcp", "model_change_set_event"),
    ("mcp", "metadata_change_set"),
    ("mcp", "metadata_change_set_event"),
    ("mcp", "tool_call_log"),
)

# Uncomment only the sheets you want to run. Dependency order is applied automatically.
MANUAL_LOAD_SELECTIONS: list[tuple[str, str]] = [
    # ("reference.xlsx", "Environment"),
    # ("reference.xlsx", "SystemType"),
    # ("reference.xlsx", "Zone"),
    # ("reference.xlsx", "ConnectionType"),
    # ("reference.xlsx", "ObjectType"),
    # ("reference.xlsx", "ConnectionParameter"),
    # ("reference.xlsx", "PurgePolicy"),
    # ("reference.xlsx", "SystemNotebook"),
    # ("reference.xlsx", "LocationType"),
    # ("reference.xlsx", "FileType"),
    # ("reference.xlsx", "Domain"),
    # ("reference.xlsx", "DataOperation"),
    # ("reference.xlsx", "ChunkType"),
    # ("reference.xlsx", "Pipeline"),
    # ("reference.xlsx", "ProcessType"),
    # ("reference.xlsx", "Currency"),
    # ("reference.xlsx", "JobType"),
    # ("reference.xlsx", "Lane"),
    # ("foundational.xlsx", "Project"),
    # ("foundational.xlsx", "System"),
    # ("foundational.xlsx", "SystemNotebookPath"),
    # ("foundational.xlsx", "Tenant"),
    # ("foundational.xlsx", "Connection"),
    # ("users_security.xlsx", "Principal"),
    # ("users_security.xlsx", "EntraPrincipalIdentity"),
    # ("users_security.xlsx", "TenantPrincipalAccess"),
    # ("operational.xlsx", "TenantMetadataDiscoveryScope"),
    # ("operational.xlsx", "ConnectionLocation"),
    # ("operational.xlsx", "ConnectionValue"),
    # ("operational.xlsx", "Object"),
    # ("operational.xlsx", "Attribute"),
    # ("operational.xlsx", "IngestionObjectMapping"),
    # ("operational.xlsx", "IngestionAttributeMapping"),
    # ("operational.xlsx", "CopyGroup"),
    # ("operational.xlsx", "MemberGroup"),
    # ("operational.xlsx", "CopyGroupControl"),
    # ("operational.xlsx", "Copy"),
    # ("operational.xlsx", "ProcessGroup"),
    # ("operational.xlsx", "Process"),
    # ("model.xlsx", "Model"),
    # ("model.xlsx", "ModelScope"),
    # ("locks.xlsx", "LockControl"),
]

_IDENTIFIER = re.compile(r"^[a-z][a-z0-9_]*$")
_ALLOWED_SCHEMAS = frozenset({"reference", "core", "security", "model"})
_ALLOWED_WORKBOOKS = frozenset(
    {
        "reference.xlsx",
        "foundational.xlsx",
        "users_security.xlsx",
        "operational.xlsx",
        "model.xlsx",
    }
)
ALLOWED_LOAD_TARGETS = {
    ("reference.xlsx", "Environment"): ("reference", "environment"),
    ("reference.xlsx", "SystemType"): ("reference", "system_type"),
    ("reference.xlsx", "Zone"): ("reference", "zone"),
    ("reference.xlsx", "ConnectionType"): ("reference", "connection_type"),
    ("reference.xlsx", "ObjectType"): ("reference", "object_type"),
    ("reference.xlsx", "ConnectionParameter"): (
        "reference",
        "connection_parameter",
    ),
    ("reference.xlsx", "PurgePolicy"): ("reference", "purge_policy"),
    ("reference.xlsx", "SystemNotebook"): ("reference", "system_notebook"),
    ("reference.xlsx", "LocationType"): ("reference", "location_type"),
    ("reference.xlsx", "FileType"): ("reference", "file_type"),
    ("reference.xlsx", "Domain"): ("reference", "domain"),
    ("reference.xlsx", "DataOperation"): ("reference", "data_operation"),
    ("reference.xlsx", "ChunkType"): ("reference", "chunk_type"),
    ("reference.xlsx", "Pipeline"): ("reference", "pipeline"),
    ("reference.xlsx", "ProcessType"): ("reference", "process_type"),
    ("reference.xlsx", "Currency"): ("reference", "currency"),
    ("reference.xlsx", "JobType"): ("reference", "job_type"),
    ("reference.xlsx", "Lane"): ("reference", "lane"),
    ("foundational.xlsx", "Project"): ("core", "project"),
    ("foundational.xlsx", "System"): ("core", "system"),
    ("foundational.xlsx", "SystemNotebookPath"): (
        "core",
        "system_notebook_path",
    ),
    ("users_security.xlsx", "Principal"): ("security", "principal"),
    ("users_security.xlsx", "EntraPrincipalIdentity"): (
        "security",
        "entra_principal_identity",
    ),
    ("users_security.xlsx", "TenantPrincipalAccess"): (
        "security",
        "tenant_principal_access",
    ),
    ("foundational.xlsx", "Tenant"): ("core", "tenant"),
    ("foundational.xlsx", "Connection"): ("core", "connection"),
    ("operational.xlsx", "TenantMetadataDiscoveryScope"): (
        "core",
        "tenant_metadata_discovery_scope",
    ),
    ("operational.xlsx", "ConnectionLocation"): (
        "core",
        "connection_location",
    ),
    ("operational.xlsx", "ConnectionValue"): ("core", "connection_value"),
    ("operational.xlsx", "Object"): ("core", "object"),
    ("operational.xlsx", "Attribute"): ("core", "attribute"),
    ("operational.xlsx", "IngestionObjectMapping"): (
        "core",
        "ingestion_object_mapping",
    ),
    ("operational.xlsx", "IngestionAttributeMapping"): (
        "core",
        "ingestion_attribute_mapping",
    ),
    ("operational.xlsx", "CopyGroup"): ("core", "copy_group"),
    ("operational.xlsx", "MemberGroup"): ("core", "member_group"),
    ("operational.xlsx", "CopyGroupControl"): (
        "core",
        "copy_group_control",
    ),
    ("operational.xlsx", "Copy"): ("core", "copy"),
    ("operational.xlsx", "ProcessGroup"): ("core", "process_group"),
    ("operational.xlsx", "Process"): ("core", "process"),
    ("model.xlsx", "Model"): ("model", "model"),
    ("model.xlsx", "ModelScope"): ("model", "model_scope"),
}


class LoaderError(ValueError):
    """A bounded failure that never includes cell values or database details."""


@dataclass(frozen=True, slots=True)
class LoadDefinition:
    workbook: str
    sheet: str
    schema: str
    table: str
    dependency_order: int
    columns: tuple[str, ...]
    required_columns: tuple[str, ...]
    source_key: tuple[str, ...]
    merge_statements: tuple[str, ...]
    case_sensitive_key_columns: tuple[str, ...] = ()
    deferred_statements: tuple[str, ...] = ()

    @property
    def target(self) -> str:
        return f"{self.schema}.{self.table}"

    @property
    def selection(self) -> tuple[str, str]:
        return self.workbook, self.sheet

    @property
    def staging_name(self) -> str:
        return f"staging_{self.schema}_{self.table}"


@dataclass(frozen=True, slots=True)
class LoaderSettings:
    database_dsn: str = field(repr=False)

    @classmethod
    def from_environment(cls, values: Mapping[str, str] | None = None) -> LoaderSettings:
        source = os.environ if values is None else values
        dsn = source.get("GDS_LOADER_DSN", "").strip()
        if not dsn:
            raise LoaderError("GDS_LOADER_DSN is required for --execute")
        if source.get("GDS_LOADER_WRITE_ACK", "") != WRITE_ACKNOWLEDGEMENT:
            raise LoaderError("GDS_LOADER_WRITE_ACK must explicitly acknowledge database writes")

        try:
            parts = conninfo_to_dict(dsn)
        except Exception as exc:
            raise LoaderError("GDS_LOADER_DSN is invalid") from exc
        user = parts.get("user")
        if (
            not isinstance(parts.get("host"), str)
            or not isinstance(parts.get("dbname"), str)
            or not isinstance(user, str)
        ):
            raise LoaderError("GDS_LOADER_DSN requires host, dbname, and user")
        if parts.get("sslmode") != "verify-full":
            raise LoaderError("GDS_LOADER_DSN requires sslmode=verify-full")
        if user.split("@", 1)[0].casefold() in {
            "gds_app_write",
            "gds_mcp_runtime",
        }:
            raise LoaderError("the MCP runtime account cannot run the Excel loader")
        return cls(database_dsn=dsn)


@dataclass(frozen=True, slots=True)
class LockActor:
    principal_id: int
    entra_tenant_id: UUID = field(repr=False)
    entra_object_id: UUID = field(repr=False)
    principal_type: str


@dataclass(frozen=True, slots=True)
class PreparedLoad:
    definition: LoadDefinition
    rows: tuple[tuple[str | None, ...], ...]


@dataclass(frozen=True, slots=True)
class PreparedLockLoad:
    rows: tuple[tuple[str, str, str, int, bool, int | None], ...]


def read_config(
    path: Path = CONFIG_PATH, *, require_complete: bool = True
) -> tuple[LoadDefinition, ...]:
    """Read and strictly validate the allowlisted merge configuration."""
    try:
        yaml = cast(_YamlModule, importlib.import_module("yaml"))
    except ImportError as exc:
        raise LoaderError("PyYAML is required; run uv sync") from exc
    try:
        raw = yaml.safe_load(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise LoaderError(f"merge configuration not found: {path.name}") from exc
    except yaml.YAMLError as exc:
        raise LoaderError("merge configuration is not valid YAML") from exc

    if not isinstance(raw, dict) or raw.get("version") != 1:
        raise LoaderError("merge configuration version must be 1")
    raw_loads = raw.get("loads")
    if not isinstance(raw_loads, list) or not raw_loads:
        raise LoaderError("merge configuration must contain loads")

    definitions: list[LoadDefinition] = []
    seen_selections: set[tuple[str, str]] = set()
    seen_orders: set[int] = set()
    for item in raw_loads:
        if not isinstance(item, dict):
            raise LoaderError("each load configuration must be an object")
        try:
            definition = LoadDefinition(
                workbook=str(item["workbook"]),
                sheet=str(item["sheet"]),
                schema=str(item["schema"]),
                table=str(item["table"]),
                dependency_order=int(item["dependency_order"]),
                columns=_string_tuple(item["columns"], "columns"),
                required_columns=_string_tuple(
                    item.get("required_columns", []),
                    "required_columns",
                    allow_empty=True,
                ),
                source_key=_string_tuple(item["source_key"], "source_key"),
                merge_statements=_string_tuple(item["merge_statements"], "merge_statements"),
                case_sensitive_key_columns=_string_tuple(
                    item.get("case_sensitive_key_columns", []),
                    "case_sensitive_key_columns",
                    allow_empty=True,
                ),
                deferred_statements=_string_tuple(
                    item.get("deferred_statements", []),
                    "deferred_statements",
                    allow_empty=True,
                ),
            )
        except (KeyError, TypeError, ValueError) as exc:
            raise LoaderError("a load configuration is incomplete") from exc
        validate_definition(definition)
        if definition.selection in seen_selections:
            raise LoaderError("workbook and sheet selections must be unique")
        if definition.dependency_order in seen_orders:
            raise LoaderError("dependency_order values must be unique")
        seen_selections.add(definition.selection)
        seen_orders.add(definition.dependency_order)
        definitions.append(definition)
    if require_complete and seen_selections != set(ALLOWED_LOAD_TARGETS):
        raise LoaderError("merge configuration must define every approved workbook sheet")
    return tuple(sorted(definitions, key=lambda value: value.dependency_order))


def _string_tuple(value: Any, name: str, *, allow_empty: bool = False) -> tuple[str, ...]:
    if (
        not isinstance(value, list)
        or (not value and not allow_empty)
        or not all(isinstance(v, str) for v in value)
    ):
        raise LoaderError(f"{name} must be a non-empty string list")
    return tuple(value)


def validate_definition(definition: LoadDefinition) -> None:
    if definition.workbook not in _ALLOWED_WORKBOOKS:
        raise LoaderError("merge configuration contains an unsupported workbook")
    if definition.schema not in _ALLOWED_SCHEMAS:
        raise LoaderError("merge configuration contains an unsupported schema")
    if not _IDENTIFIER.fullmatch(definition.table):
        raise LoaderError("merge configuration contains an invalid table")
    if definition.dependency_order < 1:
        raise LoaderError("dependency_order must be positive")
    if len(definition.sheet) > 31 or not definition.sheet.strip():
        raise LoaderError("sheet names must be 1-31 characters")
    expected_target = ALLOWED_LOAD_TARGETS.get(definition.selection)
    if expected_target != (definition.schema, definition.table):
        raise LoaderError("merge configuration contains a non-allowlisted workbook target")
    if len(set(definition.columns)) != len(definition.columns):
        raise LoaderError("configured columns must be unique")
    if any(not _IDENTIFIER.fullmatch(column) for column in definition.columns):
        raise LoaderError("configured columns must be lowercase SQL identifiers")
    if not set(definition.required_columns).issubset(definition.columns):
        raise LoaderError("required_columns must exist in columns")
    if not set(definition.source_key).issubset(definition.columns):
        raise LoaderError("source_key must exist in columns")
    if not set(definition.case_sensitive_key_columns).issubset(definition.source_key):
        raise LoaderError("case_sensitive_key_columns must exist in source_key")
    for statement in (*definition.merge_statements, *definition.deferred_statements):
        _validate_merge_statement(statement, definition)


def _validate_merge_statement(statement: str, definition: LoadDefinition) -> None:
    body = statement.strip()
    if body.endswith(";"):
        body = body[:-1]
    if ";" in body or not re.match(r"(?is)^MERGE\s+INTO\s+", body):
        raise LoaderError("each configured SQL item must be exactly one MERGE statement")
    target_pattern = rf"(?is)^MERGE\s+INTO\s+{definition.schema}\.{definition.table}\b"
    if re.match(target_pattern, body) is None:
        raise LoaderError("a MERGE statement targets the wrong configured table")
    if "{staging_table}" not in body:
        raise LoaderError("each MERGE statement must use {staging_table}")
    if re.search(r"(?is)\b(DROP|TRUNCATE|DELETE|ALTER|CREATE|GRANT|REVOKE|CALL)\b", body):
        raise LoaderError("merge configuration contains a forbidden SQL operation")


def select_loads(
    definitions: Sequence[LoadDefinition], selections: Iterable[tuple[str, str]]
) -> tuple[LoadDefinition | str, ...]:
    """Resolve explicit selections and retain canonical dependency order."""
    lookup = {definition.selection: definition for definition in definitions}
    selected: list[LoadDefinition | str] = []
    seen: set[tuple[str, str]] = set()
    for selection in selections:
        if selection == (LOCK_WORKBOOK, LOCK_SHEET):
            if selection not in seen:
                selected.append("lock")
                seen.add(selection)
            continue
        definition = lookup.get(selection)
        if definition is None:
            raise LoaderError(f"unknown workbook/sheet selection: {selection[0]} / {selection[1]}")
        if selection not in seen:
            selected.append(definition)
            seen.add(selection)
    normal = sorted(
        (item for item in selected if isinstance(item, LoadDefinition)),
        key=lambda value: value.dependency_order,
    )
    if any(item == "lock" for item in selected):
        return (*normal, "lock")
    return tuple(normal)


def prepare_selected_loads(
    selected: Sequence[LoadDefinition | str], data_dir: Path = DATA_DIR
) -> tuple[PreparedLoad | PreparedLockLoad, ...]:
    prepared: list[PreparedLoad | PreparedLockLoad] = []
    for item in selected:
        if isinstance(item, str):
            if item != "lock":
                raise LoaderError("selected load kind is invalid")
            prepared.append(PreparedLockLoad(rows=_read_lock_rows(data_dir / LOCK_WORKBOOK)))
        else:
            rows = _read_sheet_rows(data_dir / item.workbook, item.sheet, item.columns)
            _validate_rows(item, rows)
            prepared.append(PreparedLoad(definition=item, rows=rows))
    return tuple(prepared)


def _read_sheet_rows(
    workbook_path: Path, sheet_name: str, expected_columns: Sequence[str]
) -> tuple[tuple[str | None, ...], ...]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise LoaderError("openpyxl is required; run uv sync") from exc
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except FileNotFoundError as exc:
        raise LoaderError(f"workbook not found: {workbook_path.name}") from exc
    try:
        if sheet_name not in workbook.sheetnames:
            raise LoaderError(f"sheet not found: {workbook_path.name} / {sheet_name}")
        sheet = workbook[sheet_name]
        header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = tuple(_normalize_header(value) for value in header_values)
        if headers != tuple(expected_columns):
            raise LoaderError(
                f"sheet headers do not match config: {workbook_path.name} / {sheet_name}"
            )

        rows: list[tuple[str | None, ...]] = []
        for values in sheet.iter_rows(min_row=2, max_col=len(headers), values_only=True):
            converted = tuple(_cell_text(value) for value in values)
            if any(value not in (None, "") for value in converted):
                rows.append(converted)
        return tuple(rows)
    finally:
        workbook.close()


def _normalize_header(value: Any) -> str:
    return "" if value is None else str(value).strip().casefold()


def _cell_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _validate_rows(definition: LoadDefinition, rows: Sequence[tuple[str | None, ...]]) -> None:
    column_index = {column: index for index, column in enumerate(definition.columns)}
    for row in rows:
        if len(row) != len(definition.columns):
            raise LoaderError(f"row width mismatch: {definition.workbook} / {definition.sheet}")
        if any(not _nonblank(row[column_index[column]]) for column in definition.required_columns):
            raise LoaderError(
                f"required value is blank: {definition.workbook} / {definition.sheet}"
            )

    seen: set[tuple[str, ...]] = set()
    for row in rows:
        key = tuple(
            _normalize_source_key_value(
                row[column_index[column]],
                case_sensitive=column in definition.case_sensitive_key_columns,
            )
            for column in definition.source_key
        )
        if key in seen:
            raise LoaderError(f"duplicate source key: {definition.workbook} / {definition.sheet}")
        seen.add(key)


def _normalize_source_key_value(value: str | None, *, case_sensitive: bool) -> str:
    normalized = value or ""
    return normalized if case_sensitive else normalized.strip().casefold()


def _nonblank(value: str | None) -> bool:
    return value is not None and bool(value.strip())


def _read_lock_rows(
    workbook_path: Path,
) -> tuple[tuple[str, str, str, int, bool, int | None], ...]:
    raw_rows = _read_sheet_rows(workbook_path, LOCK_SHEET, LOCK_COLUMNS)
    rows: list[tuple[str, str, str, int, bool, int | None]] = []
    seen: set[tuple[str, str, str, int]] = set()
    for raw in raw_rows:
        schema = (raw[0] or "").strip().casefold()
        table = (raw[1] or "").strip().casefold()
        id_column = (raw[2] or "").strip().casefold()
        target = (schema, table, id_column)
        if target not in LOCK_TARGETS:
            raise LoaderError("LockControl contains a non-allowlisted target")
        try:
            id_value = int((raw[3] or "").strip())
        except ValueError as exc:
            raise LoaderError("LockControl id_value must be a positive integer") from exc
        if id_value < 1:
            raise LoaderError("LockControl id_value must be a positive integer")
        is_locked = _parse_lock_boolean(raw[4])
        expected_revision_text = (raw[5] or "").strip()
        expected_revision: int | None = None
        if schema == "model":
            try:
                expected_revision = int(expected_revision_text)
            except ValueError as exc:
                raise LoaderError("LockControl model rows require expected_model_revision") from exc
            if expected_revision < 1:
                raise LoaderError("LockControl model rows require expected_model_revision")
        elif expected_revision_text:
            raise LoaderError("LockControl Object rows must leave expected_model_revision blank")
        row_key = (*target, id_value)
        if row_key in seen:
            raise LoaderError("LockControl contains a duplicate target ID")
        seen.add(row_key)
        rows.append((*target, id_value, is_locked, expected_revision))
    return tuple(rows)


def _parse_lock_boolean(value: str | None) -> bool:
    normalized = (value or "").strip().casefold()
    if normalized in {"1", "true"}:
        return True
    if normalized in {"0", "false"}:
        return False
    raise LoaderError("LockControl is_locked must be 1 or 0")


def execute_prepared_loads(
    connection: Any,
    prepared: Sequence[PreparedLoad | PreparedLockLoad],
) -> None:
    """Stage and merge all selected sheets in one database transaction."""
    if bool(getattr(connection, "autocommit", False)):
        raise LoaderError("Excel loads require a non-autocommit database connection")
    has_data_load = any(isinstance(item, PreparedLoad) for item in prepared)
    has_lock_load = any(isinstance(item, PreparedLockLoad) for item in prepared)
    if has_data_load and has_lock_load:
        raise LoaderError("LockControl must run separately from data workbooks")

    deferred: list[tuple[LoadDefinition, str, int]] = []
    with connection.cursor() as cursor:
        cursor.execute("SET LOCAL lock_timeout = '5s'")
        cursor.execute("SET LOCAL statement_timeout = '5min'")
        _assert_dedicated_loader_session(cursor)
        if has_data_load:
            _assert_bootstrap_window(cursor)
            cursor.execute(
                "SET CONSTRAINTS core.uq_attribute_object_ordinal, "
                "core.uq_copy_group_order DEFERRED"
            )
        for item in prepared:
            if isinstance(item, PreparedLockLoad):
                _execute_lock_load(cursor, item)
                continue
            definition = item.definition
            create_temp_table(cursor, definition.staging_name, definition.columns)
            insert_staging_rows(cursor, definition.staging_name, definition.columns, item.rows)
            staging_table = f'pg_temp."{definition.staging_name}"'
            for statement in definition.merge_statements:
                cursor.execute(statement.replace("{staging_table}", staging_table))
                _require_complete_merge(cursor, definition, len(item.rows))
            deferred.extend(
                (definition, statement, len(item.rows))
                for statement in definition.deferred_statements
            )

        for definition, statement, expected_rows in deferred:
            staging_table = f'pg_temp."{definition.staging_name}"'
            cursor.execute(statement.replace("{staging_table}", staging_table))
            _require_complete_merge(cursor, definition, expected_rows)


def _assert_bootstrap_window(cursor: Any) -> None:
    cursor.execute(
        "SELECT pg_try_advisory_xact_lock(hashtextextended("
        "'gds_excel_bootstrap_maintenance', 0)) AS maintenance_lock"
    )
    lock_result = cursor.fetchone()
    if lock_result is None or not bool(_row_value(lock_result, 0, "maintenance_lock")):
        raise LoaderError("another Excel bootstrap load is already running")

    target_tables = sorted(set(ALLOWED_LOAD_TARGETS.values()))
    targets = ", ".join(f'"{schema}"."{table}"' for schema, table in target_tables)
    cursor.execute(f"LOCK TABLE {targets} IN SHARE ROW EXCLUSIVE MODE")
    activity_query = " UNION ALL ".join(
        f'SELECT 1 FROM "{schema}"."{table}"' for schema, table in BOOTSTRAP_ACTIVITY_TABLES
    )
    cursor.execute(f"SELECT NOT EXISTS ({activity_query}) AS bootstrap_allowed")
    result = cursor.fetchone()
    allowed = bool(_row_value(result, 0, "bootstrap_allowed")) if result is not None else False
    if not allowed:
        raise LoaderError("loader writes are allowed only before governed runtime activity")


def _assert_dedicated_loader_session(cursor: Any) -> None:
    cursor.execute(
        "SELECT NOT role_record.rolsuper "
        "AND NOT role_record.rolbypassrls "
        "AND current_user = session_user "
        "AND NOT pg_has_role(role_record.oid, database_record.datdba, 'member') "
        "AND NOT EXISTS ("
        "SELECT 1 FROM pg_class AS relation "
        "JOIN pg_namespace AS namespace ON namespace.oid = relation.relnamespace "
        "WHERE namespace.nspname IN ('reference', 'core', 'security', 'model') "
        "AND relation.relkind IN ('r', 'p') "
        "AND pg_has_role(role_record.oid, relation.relowner, 'member')"
        ") AS is_dedicated "
        "FROM pg_roles AS role_record "
        "JOIN pg_database AS database_record "
        "ON database_record.datname = current_database() "
        "WHERE role_record.rolname = session_user"
    )
    result = cursor.fetchone()
    dedicated = bool(_row_value(result, 0, "is_dedicated")) if result is not None else False
    if not dedicated:
        raise LoaderError("Excel loads require a dedicated non-owner, non-superuser login")


def _require_complete_merge(cursor: Any, definition: LoadDefinition, expected_rows: int) -> None:
    if cursor.rowcount != expected_rows:
        raise LoaderError(
            f"MERGE did not resolve every row: {definition.workbook} / {definition.sheet}"
        )


def create_temp_table(cursor: Any, table_name: str, columns: Sequence[str]) -> None:
    if not _IDENTIFIER.fullmatch(table_name) or any(
        not _IDENTIFIER.fullmatch(column) for column in columns
    ):
        raise LoaderError("unsafe staging identifier")
    column_sql = ", ".join(f'"{column}" TEXT' for column in columns)
    cursor.execute(f'CREATE TEMP TABLE "{table_name}" ({column_sql})')


def insert_staging_rows(
    cursor: Any,
    table_name: str,
    columns: Sequence[str],
    rows: Sequence[tuple[str | None, ...]],
) -> None:
    if not rows:
        return
    quoted_columns = ", ".join(f'"{column}"' for column in columns)
    placeholders = ", ".join(["%s"] * len(columns))
    cursor.executemany(
        f'INSERT INTO "{table_name}" ({quoted_columns}) VALUES ({placeholders})', rows
    )


def _resolve_lock_actor(cursor: Any) -> LockActor:
    cursor.execute(
        "SELECT principal.principal_id, identity.entra_tenant_id, "
        "identity.entra_object_id, principal.principal_type "
        "FROM security.principal AS principal "
        "JOIN security.entra_principal_identity AS identity "
        "ON identity.principal_id = principal.principal_id "
        "WHERE principal.is_active AND identity.is_active AND ("
        "(principal.principal_type = 'user' "
        "AND lower(principal.principal_email) = lower(session_user)) OR "
        "(principal.principal_type = 'service_principal' "
        "AND principal.service_principal_application_id::TEXT = session_user)"
        ")"
    )
    rows = cursor.fetchall()
    if len(rows) != 1:
        raise LoaderError("database login must map to exactly one active Entra Principal identity")
    row = rows[0]
    return LockActor(
        principal_id=int(_row_value(row, 0, "principal_id")),
        entra_tenant_id=UUID(str(_row_value(row, 1, "entra_tenant_id"))),
        entra_object_id=UUID(str(_row_value(row, 2, "entra_object_id"))),
        principal_type=str(_row_value(row, 3, "principal_type")),
    )


def _execute_lock_load(cursor: Any, load: PreparedLockLoad) -> None:
    actor = _resolve_lock_actor(cursor)
    create_temp_table(cursor, "staging_lock_control", LOCK_COLUMNS)
    staging_rows = tuple(
        (
            schema,
            table,
            id_column,
            str(id_value),
            "true" if locked else "false",
            str(expected_revision) if expected_revision is not None else None,
        )
        for schema, table, id_column, id_value, locked, expected_revision in load.rows
    )
    insert_staging_rows(cursor, "staging_lock_control", LOCK_COLUMNS, staging_rows)

    changed_model_revisions: dict[int, int] = {}
    seen_model_revisions: dict[int, int] = {}
    for (schema, table, id_column), lock_column in LOCK_TARGETS.items():
        requested = sum(
            1
            for row_schema, row_table, row_id_column, _, _, _ in load.rows
            if (row_schema, row_table, row_id_column) == (schema, table, id_column)
        )
        if requested == 0:
            continue

        policy = "tenant_metadata_write" if schema == "core" else "tenant_model_write"
        if (schema, table) == ("core", "object"):
            cursor.execute(
                'SELECT target."object_id", connection.tenant_id, '
                'target."is_locked" IS DISTINCT FROM source.is_locked::BOOLEAN AS is_changed, '
                "NULL::BIGINT AS model_id, NULL::BIGINT AS model_revision, "
                "NULL::BIGINT AS expected_model_revision "
                'FROM pg_temp."staging_lock_control" AS source '
                'JOIN "core"."object" AS target '
                'ON target."object_id" = source.id_value::BIGINT '
                'JOIN "core"."connection" AS connection '
                "ON connection.connection_id = target.connection_id "
                "WHERE source.schema_name = %s AND source.table_name = %s "
                "AND source.id_column_name = %s FOR UPDATE OF target",
                (schema, table, id_column),
            )
        else:
            cursor.execute(
                f'SELECT target."{id_column}", model_record.tenant_id, '
                f'target."{lock_column}" IS DISTINCT FROM source.is_locked::BOOLEAN AS is_changed, '
                "target.model_id, model_record.model_revision, "
                "source.expected_model_revision::BIGINT "
                'FROM pg_temp."staging_lock_control" AS source '
                f'JOIN "{schema}"."{table}" AS target '
                f'ON target."{id_column}" = source.id_value::BIGINT '
                'JOIN "model"."model" AS model_record '
                "ON model_record.model_id = target.model_id "
                "WHERE source.schema_name = %s AND source.table_name = %s "
                "AND source.id_column_name = %s FOR UPDATE OF target, model_record",
                (schema, table, id_column),
            )
        target_rows = cursor.fetchall()
        if len(target_rows) != requested:
            raise LoaderError("LockControl references an ID that does not exist")

        for row in target_rows:
            model_id_value = _row_value(row, 3, "model_id")
            if model_id_value is None:
                continue
            model_id = int(model_id_value)
            current_revision = int(_row_value(row, 4, "model_revision"))
            expected_revision = int(_row_value(row, 5, "expected_model_revision"))
            if current_revision != expected_revision:
                raise LoaderError("LockControl expected Model revision is stale")
            previous_revision = seen_model_revisions.setdefault(model_id, expected_revision)
            if previous_revision != expected_revision:
                raise LoaderError("LockControl has inconsistent expected Model revisions")

        tenant_ids = {_row_value(row, 1, "tenant_id") for row in target_rows}
        for tenant_id in tenant_ids:
            cursor.execute(
                "SELECT principal_id, authorized "
                "FROM security.authorize_tenant_operation(%s, %s, %s, %s, %s)",
                (
                    actor.entra_tenant_id,
                    actor.entra_object_id,
                    actor.principal_type,
                    tenant_id,
                    policy,
                ),
            )
            authorization = cursor.fetchone()
            if authorization is None or not bool(_row_value(authorization, 1, "authorized")):
                raise LoaderError("LockControl authorization or Tenant Lock check failed")
            if int(_row_value(authorization, 0, "principal_id")) != actor.principal_id:
                raise LoaderError("LockControl actor did not resolve consistently")

        changed_rows = [row for row in target_rows if bool(_row_value(row, 2, "is_changed"))]
        if not changed_rows:
            continue
        actor_label = f"principal:{actor.principal_id}"
        cursor.execute(
            f'UPDATE "{schema}"."{table}" AS target '
            f'SET "{lock_column}" = source.is_locked::BOOLEAN, '
            "updated_time = CURRENT_TIMESTAMP, updated_by = %s "
            'FROM pg_temp."staging_lock_control" AS source '
            f"WHERE source.schema_name = %s AND source.table_name = %s "
            f'AND source.id_column_name = %s AND target."{id_column}" = source.id_value::BIGINT '
            f'AND target."{lock_column}" IS DISTINCT FROM source.is_locked::BOOLEAN',
            (actor_label, schema, table, id_column),
        )
        if cursor.rowcount != len(changed_rows):
            raise LoaderError("LockControl target changed concurrently")
        for row in changed_rows:
            model_id_value = _row_value(row, 3, "model_id")
            if model_id_value is not None:
                changed_model_revisions[int(model_id_value)] = int(
                    _row_value(row, 5, "expected_model_revision")
                )

    if changed_model_revisions:
        actor_label = f"principal:{actor.principal_id}"
        model_ids = sorted(changed_model_revisions)
        for model_id in model_ids:
            cursor.execute(
                "UPDATE model.model SET model_revision = model_revision + 1, "
                "updated_time = CURRENT_TIMESTAMP, updated_by = %s "
                "WHERE model_id = %s AND model_revision = %s",
                (actor_label, model_id, changed_model_revisions[model_id]),
            )
            if cursor.rowcount != 1:
                raise LoaderError("LockControl Model revision changed concurrently")
        cursor.execute(
            "INSERT INTO model.model_revision_transaction "
            "(model_id, change_kind, changed_by) "
            "SELECT affected.model_id, 'excel_lock_control', %s "
            "FROM unnest(%s::BIGINT[]) AS affected(model_id)",
            (actor_label, model_ids),
        )
        if cursor.rowcount != len(model_ids):
            raise LoaderError("LockControl could not audit every affected Model")


def _row_value(row: Any, index: int, name: str) -> Any:
    if isinstance(row, Mapping):
        return row[name]
    return row[index]


def _selection_from_arguments(
    definitions: Sequence[LoadDefinition], workbook: str | None, sheet: str | None
) -> list[tuple[str, str]]:
    if sheet and not workbook:
        raise LoaderError("--sheet requires --workbook")
    if workbook is None:
        return list(MANUAL_LOAD_SELECTIONS)
    if workbook == LOCK_WORKBOOK:
        if sheet not in (None, LOCK_SHEET):
            raise LoaderError("locks.xlsx contains only LockControl")
        return [(LOCK_WORKBOOK, LOCK_SHEET)]
    if workbook not in _ALLOWED_WORKBOOKS:
        raise LoaderError("unknown workbook")
    if sheet:
        return [(workbook, sheet)]
    return [definition.selection for definition in definitions if definition.workbook == workbook]


def _print_inventory(definitions: Sequence[LoadDefinition]) -> None:
    for definition in definitions:
        print(
            f"{definition.dependency_order:03d}  {definition.workbook} / "
            f"{definition.sheet} -> {definition.target}"
        )
    print(f"999  {LOCK_WORKBOOK} / {LOCK_SHEET} -> allowlisted lock columns")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--list", action="store_true", help="list configured sheets")
    parser.add_argument(
        "--workbook", help="load one workbook; all its sheets if --sheet is omitted"
    )
    parser.add_argument("--sheet", help="load one sheet from --workbook")
    parser.add_argument(
        "--execute",
        action="store_true",
        help="connect and write; without this flag only validation runs",
    )
    return parser


def main(
    argv: Sequence[str] | None = None,
    connector: Callable[..., Any] = psycopg.connect,
) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        definitions = read_config()
        if args.list:
            _print_inventory(definitions)
            return 0
        selections = _selection_from_arguments(definitions, args.workbook, args.sheet)
        if not selections:
            raise LoaderError("uncomment a MANUAL_LOAD_SELECTIONS row or pass --workbook")
        selected = select_loads(definitions, selections)
        if any(item == "lock" for item in selected) and any(
            isinstance(item, LoadDefinition) for item in selected
        ):
            raise LoaderError("LockControl must run separately from data workbooks")
        prepared = prepare_selected_loads(selected)

        for item in prepared:
            if isinstance(item, PreparedLockLoad):
                print(f"validated {LOCK_WORKBOOK} / {LOCK_SHEET}: {len(item.rows)} row(s)")
            else:
                print(
                    f"validated {item.definition.workbook} / {item.definition.sheet} "
                    f"-> {item.definition.target}: {len(item.rows)} row(s)"
                )
        if not args.execute:
            print("dry run complete; no database connection was opened")
            return 0

        try:
            dotenv = cast(_DotenvModule, importlib.import_module("dotenv"))
            load_dotenv = dotenv.load_dotenv
        except ImportError as exc:
            raise LoaderError("python-dotenv is required; run uv sync") from exc
        load_dotenv(dotenv_path=ENV_PATH, override=False)
        settings = LoaderSettings.from_environment()
        try:
            with connector(
                settings.database_dsn,
                application_name="gds_excel_loader",
            ) as connection:
                execute_prepared_loads(connection, prepared)
        except LoaderError:
            raise
        except Exception as exc:
            raise LoaderError("database load failed; no success confirmation") from exc
        print("load committed")
        return 0
    except LoaderError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2
    except Exception:
        print("ERROR: loader input validation failed", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
